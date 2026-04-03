"""Task/Service Log module - handles task-based time tracking.

This module provides functionality for employees to log time spent on specific
tasks or services, with categorization, priority tracking, and detailed
problem/solution documentation.
"""

import logging
from datetime import date, datetime
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import (
    get_active_ts_log,
    start_ts_log,
    stop_ts_log,
    get_ts_logs_for_range,
    compute_ts_durations,
    update_ts_log_with_details,
    get_ts_log_by_id,
    update_ts_log,
)
from area_logger import login_required

logger = logging.getLogger(__name__)
ts_log_bp = Blueprint("ts_log", __name__, url_prefix="/ts-log")


@ts_log_bp.route("/", methods=["GET", "POST"])
@login_required
def index():
    """Main T/S log view - display active task log and history.
    
    Handles:
    - GET: Display current task log status and history
    - POST: Start or stop task logging
    
    Query Parameters:
        from: Start date (ISO format, defaults to today)
        to: End date (ISO format, defaults to start date)
    
    Returns:
        Rendered template with active T/S log and logs for date range.
    """
    user_id = session["user_id"]
    active_ts = get_active_ts_log(user_id)

    if request.method == "POST":
        action = request.form.get("action")

        try:
            if action == "start":
                task_name = request.form.get("task_name", "").strip()
                station = request.form.get("station", "").strip()
                problem = request.form.get("problem", "").strip()
                solution = request.form.get("solution", "").strip()
                description = request.form.get("description", "").strip()
                category = request.form.get("category", "").strip()
                priority = request.form.get("priority", "Low").strip()
                
                if not task_name:
                    flash("Task name is required to start T/S log.", "error")
                    logger.warning(f"User {user_id} attempted to start T/S log without task name")
                else:
                    if active_ts:
                        flash("You already have an active T/S log. Stop it before starting a new one.", "error")
                        logger.warning(f"User {user_id} attempted to start with active T/S log")
                    else:
                        # Start basic log first
                        start_ts_log(user_id, task_name, description, category)
                        
                        # Get the newly created log
                        new_log = get_active_ts_log(user_id)
                        
                        # Update with detailed information if provided
                        if new_log and (station or problem or solution or priority != "Low"):
                            update_ts_log_with_details(
                                new_log["id"],
                                station=station if station else None,
                                problem=problem if problem else None,
                                solution=solution if solution else None,
                                priority=priority,
                                status="In Progress"
                            )
                        
                        logger.info(f"User {user_id} started T/S log: {task_name}")
                return redirect(url_for("ts_log.index"))

            if action == "stop":
                if active_ts:
                    # Update status to Completed when stopping
                    update_ts_log_with_details(
                        active_ts["id"],
                        status="Completed"
                    )
                    stop_ts_log(active_ts["id"])
                    logger.info(f"User {user_id} stopped T/S log")
                return redirect(url_for("ts_log.index"))
        except Exception as e:
            logger.error(f"Error in T/S logger: {e}")
            flash("An error occurred. Please try again.", "error")
            return redirect(url_for("ts_log.index"))

    # Date range filters for history
    start_date = request.args.get("from") or date.today().isoformat()
    end_date = request.args.get("to") or start_date

    raw_logs = get_ts_logs_for_range(user_id, start_date, end_date)
    logs = compute_ts_durations(raw_logs)

    return render_template(
        "ts_log.html",
        active_ts=active_ts,
        logs=logs,
        start_date=start_date,
        end_date=end_date,
    )


@ts_log_bp.route("/api/log/<int:log_id>", methods=["GET"])
@login_required
def api_get_log(log_id: int):
    """API endpoint to get T/S log details (for AJAX).
    
    Args:
        log_id: T/S log ID to retrieve.
        
    Returns:
        JSON with full log details if user owns the log, 403 otherwise.
    """
    user_id = session["user_id"]
    
    try:
        log = get_ts_log_by_id(log_id)
        
        if not log or log['user_id'] != user_id:
            return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
        return jsonify({
            'status': 'success',
            'log': dict(log)
        })
    except Exception as e:
        logger.error(f"Error fetching T/S log: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@ts_log_bp.route("/api/log/<int:log_id>/update", methods=["POST"])
@login_required
def api_update_log(log_id: int):
    """API endpoint to update T/S log details (for AJAX).
    
    JSON Parameters:
        station: Station/location
        problem: Problem description
        solution: Solution summary
        status: Task status
        priority: Task priority
    
    Returns:
        JSON with status of update operation.
    """
    user_id = session["user_id"]
    
    try:
        log = get_ts_log_by_id(log_id)
        
        if not log or log['user_id'] != user_id:
            return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
        data = request.get_json()
        
        # Update the log with provided fields
        update_ts_log_with_details(
            log_id,
            station=data.get('station'),
            problem=data.get('problem'),
            solution=data.get('solution'),
            status=data.get('status'),
            priority=data.get('priority'),
        )
        
        logger.debug(f"Updated T/S log {log_id} for user {user_id}")
        
        return jsonify({'status': 'success', 'message': 'Log updated'})
        
    except Exception as e:
        logger.error(f"Error updating T/S log: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@ts_log_bp.route("/edit/<int:log_id>", methods=["GET", "POST"])
@login_required
def edit_log(log_id: int):
    """Edit a T/S log entry (only if owned by user and not active).
    
    Args:
        log_id: T/S log ID to edit.
        
    Returns:
        Edit form or redirect after update.
    """
    user_id = session["user_id"]
    
    try:
        # Get the log to edit
        log = get_ts_log_by_id(log_id)
        
        if not log or log['user_id'] != user_id:
            flash("Access denied or log not found.", "error")
            return redirect(url_for("ts_log.index"))
        
        # Check if log is active (can't edit active logs)
        if not log['end_time']:
            flash("Cannot edit active T/S logs. Stop the log first.", "error")
            return redirect(url_for("ts_log.index"))
        
        if request.method == "POST":
            # Process form submission
            task_name = request.form.get("task_name", "").strip()
            station = request.form.get("station", "").strip()
            problem = request.form.get("problem", "").strip()
            solution = request.form.get("solution", "").strip()
            description = request.form.get("description", "").strip()
            category = request.form.get("category", "").strip()
            priority = request.form.get("priority", "Low")
            status = request.form.get("status", "Pending")
            start_time = request.form.get("start_time")
            end_time = request.form.get("end_time")
            
            # Validate input
            if not task_name or not start_time or not end_time:
                flash("Task name, start time, and end time are required.", "error")
                return render_template("edit_ts_log.html", log=log)
            
            try:
                # Parse and validate times
                from datetime import datetime
                start_dt = datetime.fromisoformat(start_time.replace('T', ' '))
                end_dt = datetime.fromisoformat(end_time.replace('T', ' '))
                
                if end_dt <= start_dt:
                    flash("End time must be after start time.", "error")
                    return render_template("edit_ts_log.html", log=log)
                
                # Update the log
                success = update_ts_log(
                    log_id=log_id,
                    user_id=user_id,
                    task_name=task_name,
                    station=station,
                    problem=problem,
                    solution=solution,
                    description=description,
                    start_time=start_dt.isoformat(),
                    end_time=end_dt.isoformat(),
                    status=status,
                    priority=priority,
                    category=category
                )
                
                if success:
                    flash("T/S log updated successfully.", "success")
                    logger.info(f"User {user_id} updated T/S log {log_id}")
                    return redirect(url_for("ts_log.index"))
                else:
                    flash("Failed to update T/S log.", "error")
                    
            except ValueError as e:
                flash(f"Invalid date/time format: {e}", "error")
            except Exception as e:
                logger.error(f"Error updating T/S log {log_id}: {e}")
                flash("An error occurred while updating the log.", "error")
        
        # GET request - show edit form
        # Convert timestamps to datetime-local format for form controls
        editable_log = dict(log)
        try:
            start_dt = datetime.fromisoformat(log['start_time'])
            editable_log['start_time_local'] = start_dt.strftime('%Y-%m-%dT%H:%M')
        except Exception:
            editable_log['start_time_local'] = log['start_time'][:16] if log['start_time'] else ''

        try:
            if log['end_time']:
                end_dt = datetime.fromisoformat(log['end_time'])
                editable_log['end_time_local'] = end_dt.strftime('%Y-%m-%dT%H:%M')
                duration_hours = (end_dt - start_dt).total_seconds() / 3600.0
                editable_log['duration_hours'] = f"{duration_hours:.2f}"
            else:
                editable_log['end_time_local'] = ''
                editable_log['duration_hours'] = None
        except Exception:
            editable_log['end_time_local'] = log['end_time'][:16] if log['end_time'] else ''
            editable_log['duration_hours'] = None

        return render_template("edit_ts_log.html", log=editable_log)
        
    except Exception as e:
        logger.error(f"Error in edit_log for log {log_id}: {e}")
        flash("An error occurred.", "error")
        return redirect(url_for("ts_log.index"))


@ts_log_bp.route("/export/pdf")
@login_required
def export_pdf():
    """Export T/S logs as PDF file.
    
    Query Parameters:
        from: Start date (ISO format)
        to: End date (ISO format)
    
    Returns:
        PDF file download with formatted T/S logs report.
    """
    user_id = session["user_id"]
    start_date = request.args.get("from")
    end_date = request.args.get("to")
    
    if not start_date or not end_date:
        flash("From and To dates are required for export.", "error")
        return redirect(url_for("ts_log.index"))

    try:
        raw_logs = get_ts_logs_for_range(user_id, start_date, end_date)
        logs = compute_ts_durations(raw_logs)
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
        )
        
        # Build PDF content
        elements = []
        
        # Title
        title = Paragraph(f"T/S Logs Report - {start_date} to {end_date}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Summary
        total_hours = sum(log['duration_hours'] for log in logs)
        summary_text = f"Total Records: {len(logs)} | Total Hours: {total_hours:.2f}"
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 20))
        
        # Table data - simplified for PDF layout
        data = [['Task Name', 'Station', 'Priority', 'Status', 'Start Time', 'End Time', 'Duration (hrs)']]
        for log in logs:
            data.append([
                log['task_name'],
                log.get('station') or '',
                log.get('priority') or 'Low',
                log.get('status') or 'Pending',
                log['start_time'][:19],  # Remove microseconds
                log['end_time'][:19] if log['end_time'] else '',
                f"{log['duration_hours']:.2f}"
            ])
        
        # Create table with adjusted column widths
        table = Table(data, colWidths=[1*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1.5*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
        ]))
        
        elements.append(table)
        
        # Problem/Solution details section
        if logs:
            elements.append(Spacer(1, 20))
            details_title = Paragraph("Detailed Problem/Solution Information:", styles['Heading2'])
            elements.append(details_title)
            elements.append(Spacer(1, 12))
            
            for i, log in enumerate(logs, 1):
                detail_text = f"<b>Task {i}:</b><br/>"
                detail_text += f"<b>Problem:</b> {log.get('problem') or ''}<br/>"
                detail_text += f"<b>Solution:</b> {log.get('solution') or ''}<br/>"
                if log.get('description'):
                    detail_text += f"<b>Description:</b> {log['description']}<br/>"
                
                detail_para = Paragraph(detail_text, styles['Normal'])
                elements.append(detail_para)
                elements.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"ts_logs_{start_date}_to_{end_date}.pdf"
        logger.info(f"User {user_id} exported PDF T/S logs from {start_date} to {end_date}")
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting T/S PDF: {e}")
        flash("An error occurred during PDF export. Please try again.", "error")
        return redirect(url_for("ts_log.index"))


@ts_log_bp.route("/export/excel")
@login_required
def export_excel():
    """Export T/S logs as Excel file.
    
    Query Parameters:
        from: Start date (ISO format)
        to: End date (ISO format)
    
    Returns:
        Excel file download with formatted T/S logs.
    """
    user_id = session["user_id"]
    start_date = request.args.get("from")
    end_date = request.args.get("to")
    
    if not start_date or not end_date:
        flash("From and To dates are required for export.", "error")
        return redirect(url_for("ts_log.index"))

    try:
        raw_logs = get_ts_logs_for_range(user_id, start_date, end_date)
        logs = compute_ts_durations(raw_logs)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TS Logs"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Headers
        headers = ['Task Name', 'Station', 'Priority', 'Status', 'Start Time', 'End Time', 'Duration (hrs)', 'Problem', 'Solution', 'Description']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data rows
        for row_num, log in enumerate(logs, 2):
            ws.cell(row=row_num, column=1, value=log['task_name']).border = border
            ws.cell(row=row_num, column=2, value=log.get('station') or '').border = border
            ws.cell(row=row_num, column=3, value=log.get('priority') or 'Low').border = border
            ws.cell(row=row_num, column=4, value=log.get('status') or 'Pending').border = border
            ws.cell(row=row_num, column=5, value=log['start_time'][:19]).border = border
            ws.cell(row=row_num, column=6, value=log['end_time'][:19] if log['end_time'] else '').border = border
            duration_cell = ws.cell(row=row_num, column=7, value=round(log['duration_hours'], 2))
            duration_cell.border = border
            duration_cell.number_format = '0.00'
            
            # Problem, Solution, Description with text wrapping
            problem_cell = ws.cell(row=row_num, column=8, value=log.get('problem') or '')
            problem_cell.border = border
            problem_cell.alignment = wrap_align
            
            solution_cell = ws.cell(row=row_num, column=9, value=log.get('solution') or '')
            solution_cell.border = border
            solution_cell.alignment = wrap_align
            
            desc_cell = ws.cell(row=row_num, column=10, value=log.get('description', ''))
            desc_cell.border = border
            desc_cell.alignment = wrap_align
        
        # Auto-adjust column widths (with max width limits for text columns)
        column_widths = [12, 15, 10, 10, 18, 18, 12, 30, 30, 25]  # Max widths
        for col_num, max_width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max_width
        
        # Summary row
        summary_row = len(logs) + 3
        ws.cell(row=summary_row, column=1, value="Total Records:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(logs))
        ws.cell(row=summary_row + 1, column=1, value="Total Hours:").font = Font(bold=True)
        total_hours = sum(log['duration_hours'] for log in logs)
        ws.cell(row=summary_row + 1, column=2, value=round(total_hours, 2))
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"ts_logs_{start_date}_to_{end_date}.xlsx"
        logger.info(f"User {user_id} exported Excel T/S logs from {start_date} to {end_date}")
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting T/S Excel: {e}")
        flash("An error occurred during Excel export. Please try again.", "error")
        return redirect(url_for("ts_log.index"))

