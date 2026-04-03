"""Area Logger module - handles time tracking by work area/factory.

This module provides functionality for employees to log time spent in different
work areas and factories, with automatic tracking of duration and charge codes.
"""

import logging
from datetime import date
import csv
from io import StringIO, BytesIO
from functools import wraps
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import (
    get_areas_with_charge_codes,
    get_active_log,
    start_logging,
    stop_logging,
    get_logs_for_range,
    compute_durations,
    get_time_log_by_id,
    update_time_log,
)

logger = logging.getLogger(__name__)
area_logger_bp = Blueprint("area_logger", __name__, url_prefix="/area-logger")


def login_required(fn):
    """Decorator to require user login for a route.
    
    Args:
        fn: View function to decorate.
        
    Returns:
        Decorated function that redirects to login if not authenticated.
    """
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("auth.login"))
        return fn(*args, **kwargs)
    return wrapper


@area_logger_bp.route("/", methods=["GET", "POST"])
@login_required
def index():
    """Main area logger view - display active log and history.
    
    Handles:
    - GET: Display current logging status and history
    - POST: Start, stop, or switch area logging
    
    Returns:
        Rendered template with areas, active log, and logs for date range.
    """
    user_id = session["user_id"]
    areas = get_areas_with_charge_codes()
    active_log = get_active_log(user_id)

    if request.method == "POST":
        action = request.form.get("action")
        area_id = request.form.get("area_id")

        try:
            if action == "start":
                if active_log:
                    flash("You already have an active log. Stop or switch first.", "error")
                    logger.warning(f"User {user_id} attempted to start with active log")
                elif area_id:
                    start_logging(user_id, int(area_id))
                    logger.info(f"User {user_id} started logging in area {area_id}")
                return redirect(url_for("area_logger.index"))

            if action == "stop":
                if active_log:
                    stop_logging(active_log["id"])
                    logger.info(f"User {user_id} stopped logging")
                return redirect(url_for("area_logger.index"))

            if action == "switch":
                if active_log:
                    stop_logging(active_log["id"])
                if area_id:
                    start_logging(user_id, int(area_id))
                    logger.info(f"User {user_id} switched to area {area_id}")
                return redirect(url_for("area_logger.index"))
        except Exception as e:
            logger.error(f"Error in area logger: {e}")
            flash("An error occurred. Please try again.", "error")
            return redirect(url_for("area_logger.index"))

    # Date range filters
    start_date = request.args.get("from") or date.today().isoformat()
    end_date = request.args.get("to") or start_date

    raw_logs = get_logs_for_range(user_id, start_date, end_date)
    logs = compute_durations(raw_logs)

    return render_template(
        "area_logger.html",
        areas=areas,
        active_log=active_log,
        logs=logs,
        start_date=start_date,
        end_date=end_date,
    )


@area_logger_bp.route("/export")
@login_required
def export_csv():
    """Export time logs as CSV file.
    
    Query Parameters:
        from: Start date (ISO format)
        to: End date (ISO format)
    
    Returns:
        CSV file download with time logs for the date range.
    """
    user_id = session["user_id"]
    start_date = request.args.get("from")
    end_date = request.args.get("to")
    
    if not start_date or not end_date:
        flash("From and To dates are required for export.", "error")
        return redirect(url_for("area_logger.index"))

    try:
        logs = compute_durations(get_logs_for_range(user_id, start_date, end_date))

        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(["Area", "Department", "Charge Code", "Start Time", "End Time", "Duration (hours)"])
        for row in logs:
            writer.writerow([
                row["area_name"],
                row.get("department_code", ""),
                row["charge_code"],
                row["start_time"],
                row["end_time"] or "",
                f"{row['duration_hours']:.3f}",
            ])

        output = si.getvalue()
        filename = f"time_logs_{start_date}_to_{end_date}.csv"
        logger.info(f"User {user_id} exported logs from {start_date} to {end_date}")
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting CSV: {e}")
        flash("An error occurred during export. Please try again.", "error")
        return redirect(url_for("area_logger.index"))


@area_logger_bp.route("/edit/<int:log_id>", methods=["GET", "POST"])
@login_required
def edit_log(log_id: int):
    """Edit a time log entry (only if owned by user and not active).
    
    Args:
        log_id: Time log ID to edit.
        
    Returns:
        Edit form or redirect after update.
    """
    user_id = session["user_id"]
    
    try:
        # Get the log to edit
        log = get_time_log_by_id(log_id)
        
        if not log or log['user_id'] != user_id:
            flash("Access denied or log not found.", "error")
            return redirect(url_for("area_logger.index"))
        
        # Check if log is active (can't edit active logs)
        if not log['end_time']:
            flash("Cannot edit active time logs. Stop the log first.", "error")
            return redirect(url_for("area_logger.index"))
        
        # Get areas for dropdown
        areas = get_areas_with_charge_codes()
        
        if request.method == "POST":
            # Process form submission
            area_id = request.form.get("area_id")
            start_time = request.form.get("start_time")
            end_time = request.form.get("end_time")
            
            # Validate input
            if not area_id or not start_time or not end_time:
                flash("Area, start time, and end time are required.", "error")
                return render_template("edit_time_log.html", log=log, areas=areas)
            
            try:
                # Parse and validate times
                from datetime import datetime
                start_dt = datetime.fromisoformat(start_time.replace('T', ' '))
                end_dt = datetime.fromisoformat(end_time.replace('T', ' '))
                
                if end_dt <= start_dt:
                    flash("End time must be after start time.", "error")
                    return render_template("edit_time_log.html", log=log, areas=areas)
                
                # Update the log
                success = update_time_log(
                    log_id=log_id,
                    user_id=user_id,
                    area_id=int(area_id),
                    start_time=start_dt.isoformat(),
                    end_time=end_dt.isoformat()
                )
                
                if success:
                    flash("Time log updated successfully.", "success")
                    logger.info(f"User {user_id} updated time log {log_id}")
                    return redirect(url_for("area_logger.index"))
                else:
                    flash("Failed to update time log.", "error")
                    
            except ValueError as e:
                flash(f"Invalid date/time format: {e}", "error")
            except Exception as e:
                logger.error(f"Error updating time log {log_id}: {e}")
                flash("An error occurred while updating the log.", "error")
        
        # GET request - show edit form
        return render_template("edit_time_log.html", log=log, areas=areas)
        
    except Exception as e:
        logger.error(f"Error in edit_log for log {log_id}: {e}")
        flash("An error occurred.", "error")
        return redirect(url_for("area_logger.index"))


@area_logger_bp.route("/export/pdf")
@login_required
def export_pdf():
    """Export time logs as PDF file.
    
    Query Parameters:
        from: Start date (ISO format)
        to: End date (ISO format)
    
    Returns:
        PDF file download with formatted time logs report.
    """
    user_id = session["user_id"]
    start_date = request.args.get("from")
    end_date = request.args.get("to")
    
    if not start_date or not end_date:
        flash("From and To dates are required for export.", "error")
        return redirect(url_for("area_logger.index"))

    try:
        logs = compute_durations(get_logs_for_range(user_id, start_date, end_date))
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
        )
        
        # Build PDF content
        elements = []
        
        # Title
        title = Paragraph(f"Time Logs Report - {start_date} to {end_date}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Summary
        total_hours = sum(log['duration_hours'] for log in logs)
        summary_text = f"Total Records: {len(logs)} | Total Hours: {total_hours:.2f}"
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Area', 'Department', 'Charge Code', 'Start Time', 'End Time', 'Duration (hrs)']]
        for log in logs:
            data.append([
                log['area_name'],
                log.get('department_code', ''),
                log['charge_code'],
                log['start_time'][:19],  # Remove microseconds
                log['end_time'][:19] if log['end_time'] else '',
                f"{log['duration_hours']:.2f}"
            ])
        
        # Create table
        table = Table(data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.8*inch, 1.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"time_logs_{start_date}_to_{end_date}.pdf"
        logger.info(f"User {user_id} exported PDF logs from {start_date} to {end_date}")
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting PDF: {e}")
        flash("An error occurred during PDF export. Please try again.", "error")
        return redirect(url_for("area_logger.index"))


@area_logger_bp.route("/export/excel")
@login_required
def export_excel():
    """Export time logs as Excel file.
    
    Query Parameters:
        from: Start date (ISO format)
        to: End date (ISO format)
    
    Returns:
        Excel file download with formatted time logs.
    """
    user_id = session["user_id"]
    start_date = request.args.get("from")
    end_date = request.args.get("to")
    
    if not start_date or not end_date:
        flash("From and To dates are required for export.", "error")
        return redirect(url_for("area_logger.index"))

    try:
        logs = compute_durations(get_logs_for_range(user_id, start_date, end_date))
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Time Logs"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        
        # Headers
        headers = ['Area', 'Department', 'Charge Code', 'Start Time', 'End Time', 'Duration (hrs)']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data rows
        for row_num, log in enumerate(logs, 2):
            ws.cell(row=row_num, column=1, value=log['area_name']).border = border
            ws.cell(row=row_num, column=2, value=log.get('department_code', '')).border = border
            ws.cell(row=row_num, column=3, value=log['charge_code']).border = border
            ws.cell(row=row_num, column=4, value=log['start_time'][:19]).border = border
            ws.cell(row=row_num, column=5, value=log['end_time'][:19] if log['end_time'] else '').border = border
            duration_cell = ws.cell(row=row_num, column=6, value=round(log['duration_hours'], 2))
            duration_cell.border = border
            duration_cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Summary row
        summary_row = len(logs) + 3
        ws.cell(row=summary_row, column=1, value="Total Records:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(logs))
        ws.cell(row=summary_row + 1, column=1, value="Total Hours:").font = Font(bold=True)
        total_hours = sum(log['duration_hours'] for log in logs)
        ws.cell(row=summary_row + 1, column=2, value=round(total_hours, 2))
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"time_logs_{start_date}_to_{end_date}.xlsx"
        logger.info(f"User {user_id} exported Excel logs from {start_date} to {end_date}")
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        flash("An error occurred during Excel export. Please try again.", "error")
        return redirect(url_for("area_logger.index"))
