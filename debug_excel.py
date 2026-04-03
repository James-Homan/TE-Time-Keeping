#!/usr/bin/env python3
"""Debug Excel export issue."""

import APP
from datetime import date
from io import BytesIO
from models import get_ts_logs_for_range, compute_ts_durations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = APP.create_app()

with app.app_context():
    # Get some test data
    user_id = 8  # From previous test
    start_date = date.today().isoformat()
    end_date = date.today().isoformat()
    
    print(f"Start date: {start_date} (type: {type(start_date)})")
    print(f"End date: {end_date} (type: {type(end_date)})")
    
    raw_logs = get_ts_logs_for_range(user_id, start_date, end_date)
    logs = compute_ts_durations(raw_logs)
    print(f"Got {len(logs)} logs")
    
    try:
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "T/S Logs"
        print(f"Sheet title set to: '{ws.title}'")
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Headers
        headers = ['Task Type', 'Station', 'Priority', 'Status', 'Start Time', 'End Time', 'Duration (hrs)', 'Problem', 'Solution', 'Description']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        print(f"Headers added: {len(headers)} columns")
        
        # Data rows
        for row_num, log in enumerate(logs, 2):
            ws.cell(row=row_num, column=1, value=log['task_name']).border = border
            
        print(f"Data rows added: {len(logs)} rows")
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        print(f"✓ Excel successfully saved: {len(buffer.getvalue())} bytes")
        
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()
